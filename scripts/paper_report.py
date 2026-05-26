"""Generiert Excel + HTML Reports aus Paper-Trading-Daten.

Wird automatisch von paper_trading.py aufgerufen — kann auch standalone:
    python -m scripts.paper_report
"""
import base64
import io
import json
from datetime import datetime, timezone
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

STATE_DIR = Path("results")
STATE_FILE = STATE_DIR / "paper_state.json"
TRADES_FILE = STATE_DIR / "paper_trades.csv"
EQUITY_FILE = STATE_DIR / "paper_equity.csv"
EXCEL_FILE = STATE_DIR / "paper_report.xlsx"
HTML_FILE = STATE_DIR / "paper_report.html"


def compute_kpis(state, equity_df, trades_df):
    initial = state["initial_capital"]
    if equity_df is None or len(equity_df) == 0:
        current = initial
    else:
        current = equity_df.iloc[-1]["total_equity"]
    total_ret = current / initial - 1
    n_days = len(equity_df) if equity_df is not None else 0
    cagr = ((current / initial) ** (252 / max(n_days, 1)) - 1) if n_days > 1 else 0
    if equity_df is not None and len(equity_df) > 1:
        eq = equity_df["total_equity"]
        max_dd = (eq / eq.cummax() - 1).min()
    else:
        max_dd = 0
    n_trades = len(trades_df) if trades_df is not None else 0
    if trades_df is not None and n_trades > 0:
        wins = (trades_df["pnl_chf"] > 0).sum()
        wr = wins / n_trades
        total_pnl = trades_df["pnl_chf"].sum()
    else:
        wins = 0; wr = 0; total_pnl = 0
    return {
        "initial_capital": initial,
        "current_equity": current,
        "total_return_pct": total_ret * 100,
        "cagr_pct": cagr * 100,
        "max_drawdown_pct": max_dd * 100,
        "n_days": n_days,
        "n_trades": n_trades,
        "win_rate_pct": wr * 100,
        "total_pnl": total_pnl,
        "started_at": state.get("started_at", "")[:10],
        "last_run": state.get("last_run", "nie"),
    }


def make_equity_chart_png_b64(equity_df, initial):
    """Equity-Kurve als base64-PNG fuer HTML-Embed."""
    if equity_df is None or len(equity_df) < 2:
        return None
    fig, ax = plt.subplots(figsize=(10, 4))
    eq = equity_df.copy()
    eq["date"] = pd.to_datetime(eq["date"])
    ax.plot(eq["date"], eq["total_equity"], linewidth=2, color="#2c7be5")
    ax.axhline(initial, color="gray", linestyle="--", alpha=0.5, label=f"Start {initial:.0f} CHF")
    ax.fill_between(eq["date"], initial, eq["total_equity"],
                    where=(eq["total_equity"] >= initial), alpha=0.15, color="green")
    ax.fill_between(eq["date"], initial, eq["total_equity"],
                    where=(eq["total_equity"] < initial), alpha=0.15, color="red")
    ax.set_xlabel("Datum")
    ax.set_ylabel("Equity (CHF)")
    ax.set_title("Equity-Kurve Paper-Trading")
    ax.grid(alpha=0.3)
    ax.legend()
    fig.autofmt_xdate()
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=80)
    plt.close(fig)
    return base64.b64encode(buf.getvalue()).decode()


def make_trades_chart_png_b64(trades_df):
    """Bar-Chart PnL pro Trade."""
    if trades_df is None or len(trades_df) == 0:
        return None
    fig, ax = plt.subplots(figsize=(10, 3))
    colors = ["green" if p > 0 else "red" for p in trades_df["pnl_chf"]]
    ax.bar(range(len(trades_df)), trades_df["pnl_chf"], color=colors, alpha=0.7)
    ax.axhline(0, color="black", linewidth=0.5)
    ax.set_xlabel("Trade #")
    ax.set_ylabel("PnL (CHF)")
    ax.set_title("PnL pro Trade")
    ax.grid(alpha=0.3)
    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=80)
    plt.close(fig)
    return base64.b64encode(buf.getvalue()).decode()


def write_excel(state, equity_df, trades_df, kpis):
    wb = Workbook()
    # --- Sheet 1: Übersicht ---
    ws = wb.active
    ws.title = "Übersicht"
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2c7be5")
    cell_bold = Font(bold=True)

    ws["A1"] = "Paper-Trading Report"
    ws["A1"].font = Font(bold=True, size=18)
    ws.merge_cells("A1:D1")

    ws["A3"] = "Generiert:"
    ws["B3"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws["A4"] = "Gestartet:"
    ws["B4"] = kpis["started_at"]
    ws["A5"] = "Letzter Bot-Run:"
    ws["B5"] = kpis["last_run"]
    for r in range(3, 6):
        ws.cell(row=r, column=1).font = cell_bold

    ws["A7"] = "KENNZAHL"
    ws["B7"] = "WERT"
    for c in ["A7", "B7"]:
        ws[c].font = header_font
        ws[c].fill = header_fill

    kpi_rows = [
        ("Startkapital", f"{kpis['initial_capital']:.2f} CHF"),
        ("Aktuelle Equity", f"{kpis['current_equity']:.2f} CHF"),
        ("Total Return", f"{kpis['total_return_pct']:+.2f} %"),
        ("CAGR (annualisiert)", f"{kpis['cagr_pct']:+.2f} %"),
        ("Max Drawdown", f"{kpis['max_drawdown_pct']:+.2f} %"),
        ("Trading-Tage", str(kpis["n_days"])),
        ("Trades gesamt", str(kpis["n_trades"])),
        ("Win-Rate", f"{kpis['win_rate_pct']:.1f} %"),
        ("Total PnL", f"{kpis['total_pnl']:+.2f} CHF"),
    ]
    for i, (label, value) in enumerate(kpi_rows, start=8):
        ws.cell(row=i, column=1, value=label).font = cell_bold
        ws.cell(row=i, column=2, value=value)

    # Tranchen
    ws["A20"] = "TRANCHEN-STATUS"
    ws["A20"].font = Font(bold=True, size=12)
    headers = ["Name", "Kapital (CHF)", "Anteil", "Hebel", "Offene Positionen"]
    for i, h in enumerate(headers):
        c = ws.cell(row=21, column=i+1, value=h)
        c.font = header_font
        c.fill = header_fill
    row = 22
    for tc_name, tc in state["tranches"].items():
        ws.cell(row=row, column=1, value=tc_name)
        ws.cell(row=row, column=2, value=round(tc["capital"], 2))
        ws.cell(row=row, column=3, value=f"{tc['capital_share']*100:.0f} %")
        ws.cell(row=row, column=4, value=f"{tc['leverage']:.1f}x")
        ws.cell(row=row, column=5, value=len(tc["open_positions"]))
        row += 1

    # Spalten-Breite
    for col_letter, width in [("A", 22), ("B", 22), ("C", 14), ("D", 12), ("E", 18)]:
        ws.column_dimensions[col_letter].width = width

    # --- Sheet 2: Trades ---
    ws2 = wb.create_sheet("Trades")
    if trades_df is not None and len(trades_df) > 0:
        for c_idx, col in enumerate(trades_df.columns, start=1):
            cell = ws2.cell(row=1, column=c_idx, value=col)
            cell.font = header_font
            cell.fill = header_fill
        for r_idx, (_, row) in enumerate(trades_df.iterrows(), start=2):
            for c_idx, val in enumerate(row, start=1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                if isinstance(val, (int, float)) and "pnl_chf" in trades_df.columns[c_idx-1]:
                    cell.font = Font(color="00C853" if val > 0 else "F44336")
        for col_letter in "ABCDEFGHIJK":
            ws2.column_dimensions[col_letter].width = 14
    else:
        ws2["A1"] = "Noch keine Trades."

    # --- Sheet 3: Equity-Kurve ---
    ws3 = wb.create_sheet("Equity-Kurve")
    if equity_df is not None and len(equity_df) > 0:
        ws3["A1"] = "Datum"; ws3["B1"] = "Total Equity"
        for c in ["A1", "B1"]:
            ws3[c].font = header_font
            ws3[c].fill = header_fill
        for r_idx, (_, row) in enumerate(equity_df.iterrows(), start=2):
            ws3.cell(row=r_idx, column=1, value=str(row["date"])[:10])
            ws3.cell(row=r_idx, column=2, value=float(row["total_equity"]))
        # Chart einbauen
        if len(equity_df) >= 2:
            chart = LineChart()
            chart.title = "Equity-Kurve"
            chart.x_axis.title = "Datum"
            chart.y_axis.title = "CHF"
            data = Reference(ws3, min_col=2, min_row=1, max_row=len(equity_df)+1)
            cats = Reference(ws3, min_col=1, min_row=2, max_row=len(equity_df)+1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10; chart.width = 20
            ws3.add_chart(chart, "D2")
        ws3.column_dimensions["A"].width = 14
        ws3.column_dimensions["B"].width = 14
    else:
        ws3["A1"] = "Noch keine Equity-Daten."

    # --- Sheet 4: Offene Positionen ---
    ws4 = wb.create_sheet("Offene Positionen")
    headers = ["Tranche", "Pair", "Richtung", "Entry-Datum", "Entry-Preis", "Position %"]
    for i, h in enumerate(headers):
        c = ws4.cell(row=1, column=i+1, value=h)
        c.font = header_font
        c.fill = header_fill
    row = 2
    for tc_name, tc in state["tranches"].items():
        for sym, pos in tc["open_positions"].items():
            ws4.cell(row=row, column=1, value=tc_name)
            ws4.cell(row=row, column=2, value=sym)
            ws4.cell(row=row, column=3, value="LONG" if pos["direction"] == 1 else "SHORT")
            ws4.cell(row=row, column=4, value=pos["entry_date"])
            ws4.cell(row=row, column=5, value=pos["entry_price"])
            ws4.cell(row=row, column=6, value=f"{pos['position_frac']*100:.0f} %")
            row += 1
    if row == 2:
        ws4.cell(row=2, column=1, value="Aktuell keine offenen Positionen.")
    for col_letter in "ABCDEF":
        ws4.column_dimensions[col_letter].width = 16

    wb.save(EXCEL_FILE)


def write_html(state, equity_df, trades_df, kpis):
    eq_chart_b64 = make_equity_chart_png_b64(equity_df, kpis["initial_capital"])
    trades_chart_b64 = make_trades_chart_png_b64(trades_df)

    # Color für Total Return
    ret_color = "#00c853" if kpis["total_return_pct"] >= 0 else "#f44336"
    pnl_color = "#00c853" if kpis["total_pnl"] >= 0 else "#f44336"

    open_pos_rows = ""
    for tc_name, tc in state["tranches"].items():
        for sym, pos in tc["open_positions"].items():
            d = "LONG" if pos["direction"] == 1 else "SHORT"
            d_color = "#00c853" if pos["direction"] == 1 else "#f44336"
            open_pos_rows += f"""
            <tr>
              <td>{tc_name}</td>
              <td><b>{sym}</b></td>
              <td style="color:{d_color}"><b>{d}</b></td>
              <td>{pos['entry_date']}</td>
              <td>{pos['entry_price']:.5f}</td>
              <td>{pos['position_frac']*100:.0f}%</td>
            </tr>"""
    if not open_pos_rows:
        open_pos_rows = "<tr><td colspan='6' style='text-align:center;color:gray'>Keine offenen Positionen</td></tr>"

    tranche_rows = ""
    for tc_name, tc in state["tranches"].items():
        start_cap = state["initial_capital"] * tc["capital_share"]
        ret = tc["capital"] / start_cap - 1
        r_color = "#00c853" if ret >= 0 else "#f44336"
        tranche_rows += f"""
        <tr>
          <td><b>{tc_name}</b></td>
          <td>{tc['capital']:.2f} CHF</td>
          <td>{tc['capital_share']*100:.0f}%</td>
          <td>{tc['leverage']:.1f}x</td>
          <td style="color:{r_color}"><b>{ret*100:+.2f}%</b></td>
          <td>{len(tc['open_positions'])}</td>
        </tr>"""

    # Letzte 20 Trades
    trades_rows = ""
    if trades_df is not None and len(trades_df) > 0:
        recent = trades_df.tail(20).iloc[::-1]
        for _, t in recent.iterrows():
            pnl = float(t["pnl_chf"])
            c = "#00c853" if pnl > 0 else "#f44336"
            trades_rows += f"""
            <tr>
              <td>{t['date']}</td>
              <td>{t['tranche']}</td>
              <td>{t['pair']}</td>
              <td>{t['direction']}</td>
              <td>{t['gross_return_pct']:+.2f}%</td>
              <td style="color:{c}"><b>{t['net_return_pct']:+.2f}%</b></td>
              <td style="color:{c}"><b>{pnl:+.2f}</b></td>
            </tr>"""
    else:
        trades_rows = "<tr><td colspan='7' style='text-align:center;color:gray'>Noch keine Trades</td></tr>"

    eq_img = f'<img src="data:image/png;base64,{eq_chart_b64}" style="width:100%;max-width:1000px"/>' if eq_chart_b64 else "<p style='color:gray'>Noch keine Equity-Kurve (warte auf 2+ Datenpunkte)</p>"
    tr_img = f'<img src="data:image/png;base64,{trades_chart_b64}" style="width:100%;max-width:1000px"/>' if trades_chart_b64 else ""

    html = f"""<!DOCTYPE html>
<html lang="de"><head>
<meta charset="UTF-8">
<title>Paper-Trading Report</title>
<style>
  body {{ font-family: -apple-system, BlinkMacSystemFont, sans-serif; max-width: 1100px; margin: 30px auto; padding: 0 20px; color: #333; background: #f5f7fa; }}
  h1 {{ color: #2c7be5; border-bottom: 3px solid #2c7be5; padding-bottom: 10px; }}
  h2 {{ color: #2c7be5; margin-top: 35px; }}
  .kpi-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin: 20px 0; }}
  .kpi {{ background: white; padding: 18px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.05); }}
  .kpi-label {{ color: #888; font-size: 12px; text-transform: uppercase; letter-spacing: 0.5px; }}
  .kpi-value {{ font-size: 24px; font-weight: bold; margin-top: 4px; }}
  table {{ width: 100%; border-collapse: collapse; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 4px rgba(0,0,0,0.05); }}
  th {{ background: #2c7be5; color: white; padding: 12px; text-align: left; font-weight: 600; }}
  td {{ padding: 10px 12px; border-bottom: 1px solid #eee; }}
  tr:last-child td {{ border-bottom: none; }}
  tr:hover {{ background: #f9fafb; }}
  .meta {{ color: #888; font-size: 13px; margin-bottom: 20px; }}
  .chart-box {{ background: white; padding: 20px; border-radius: 8px; margin: 20px 0; box-shadow: 0 2px 4px rgba(0,0,0,0.05); text-align: center; }}
</style>
</head><body>

<h1>📊 Paper-Trading Report</h1>
<div class="meta">
  Gestartet: <b>{kpis['started_at']}</b> &nbsp;|&nbsp;
  Letzter Bot-Run: <b>{kpis['last_run']}</b> &nbsp;|&nbsp;
  Generiert: <b>{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}</b>
</div>

<div class="kpi-grid">
  <div class="kpi">
    <div class="kpi-label">Aktuelle Equity</div>
    <div class="kpi-value">{kpis['current_equity']:,.2f} CHF</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">Total Return</div>
    <div class="kpi-value" style="color:{ret_color}">{kpis['total_return_pct']:+.2f}%</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">CAGR (annualisiert)</div>
    <div class="kpi-value" style="color:{ret_color}">{kpis['cagr_pct']:+.2f}%</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">Max Drawdown</div>
    <div class="kpi-value" style="color:#f44336">{kpis['max_drawdown_pct']:+.2f}%</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">Trades</div>
    <div class="kpi-value">{kpis['n_trades']}</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">Win-Rate</div>
    <div class="kpi-value">{kpis['win_rate_pct']:.1f}%</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">Total PnL</div>
    <div class="kpi-value" style="color:{pnl_color}">{kpis['total_pnl']:+.2f} CHF</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">Trading-Tage</div>
    <div class="kpi-value">{kpis['n_days']}</div>
  </div>
</div>

<h2>📈 Equity-Kurve</h2>
<div class="chart-box">{eq_img}</div>

<h2>💰 Tranchen-Status</h2>
<table>
<thead><tr><th>Name</th><th>Aktuelle Kapital</th><th>Anteil</th><th>Hebel</th><th>Return</th><th>Offen</th></tr></thead>
<tbody>{tranche_rows}</tbody>
</table>

<h2>🎯 Offene Positionen</h2>
<table>
<thead><tr><th>Tranche</th><th>Pair</th><th>Richtung</th><th>Entry-Datum</th><th>Entry-Preis</th><th>Position</th></tr></thead>
<tbody>{open_pos_rows}</tbody>
</table>

<h2>📋 Letzte 20 Trades</h2>
<table>
<thead><tr><th>Datum</th><th>Tranche</th><th>Pair</th><th>Richtung</th><th>Brutto</th><th>Netto</th><th>PnL (CHF)</th></tr></thead>
<tbody>{trades_rows}</tbody>
</table>

{f'<h2>📊 PnL pro Trade</h2><div class="chart-box">{tr_img}</div>' if tr_img else ""}

<div class="meta" style="margin-top:40px;text-align:center">
  Datenquelle: yfinance &nbsp;|&nbsp; Strategie: Williams %R Mean-Reversion + Barbell-Allokation<br>
  Dateien: results/paper_state.json, paper_trades.csv, paper_equity.csv
</div>

</body></html>"""

    HTML_FILE.write_text(html, encoding="utf-8")


def main():
    if not STATE_FILE.exists():
        print("Kein State gefunden. Starte zuerst: python -m scripts.paper_trading")
        return
    with open(STATE_FILE) as f:
        state = json.load(f)

    equity_df = pd.read_csv(EQUITY_FILE) if EQUITY_FILE.exists() else None
    trades_df = pd.read_csv(TRADES_FILE) if TRADES_FILE.exists() else None
    if equity_df is not None:
        equity_df["date"] = pd.to_datetime(equity_df["date"])

    kpis = compute_kpis(state, equity_df, trades_df)
    write_excel(state, equity_df, trades_df, kpis)
    write_html(state, equity_df, trades_df, kpis)
    print(f"  📊 Excel:  {EXCEL_FILE.absolute()}")
    print(f"  📊 HTML:   {HTML_FILE.absolute()}")


if __name__ == "__main__":
    main()
